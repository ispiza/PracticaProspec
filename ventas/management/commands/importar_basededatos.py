# ventas/management/commands/importar_basededatos.py

from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from datetime import datetime, timedelta
from decimal import Decimal

from catalogos.models import Categoria, Sucursal, MetodoPago, Cliente, Vendedor, Producto
from ventas.models import Venta, DetalleVenta


class Command(BaseCommand):
    help = 'Importa datos desde el archivo Excel DashBoard2021.xlsx'

    def add_arguments(self, parser):
        parser.add_argument('--archivo', type=str, required=True)

    def handle(self, *args, **options):
        archivo = options['archivo']
        wb = load_workbook(archivo, data_only=True)

        if 'BaseDeDatos' not in wb.sheetnames:
            self.stdout.write(self.style.ERROR("No existe la hoja 'BaseDeDatos'"))
            return

        ws = wb['BaseDeDatos']
        filas = list(ws.iter_rows(values_only=True))

        encabezados = filas[1]
        datos = filas[2:]

        ventas_temporales = {}

        for fila in datos:
            if all(v is None for v in fila):
                continue

            registro = dict(zip(encabezados, fila))

            folio            = str(registro.get('Documento', '') or '').strip()
            cliente_nombre   = str(registro.get('Cliente', '') or '').strip()
            ciudad           = str(registro.get('Ciudad', '') or '').strip()
            provincia        = str(registro.get('Provincia', '') or '').strip()
            vendedor_nombre  = str(registro.get('Vendedor', '') or '').strip()
            empresa_nombre   = str(registro.get('Empresa', '') or '').strip()
            metodo_nombre    = str(registro.get('Forma de pago', '') or '').strip()
            nombre_producto  = str(registro.get('Producto', '') or '').strip()
            categoria_nombre = str(registro.get('Categoría', '') or '').strip()

            # Manejo de fecha con soporte para número serial de Excel
            fecha_valor = registro.get('Fecha')
            if isinstance(fecha_valor, datetime):
                fecha = fecha_valor.date()
            elif isinstance(fecha_valor, (int, float)):
                fecha = (datetime(1899, 12, 30) + timedelta(days=int(fecha_valor))).date()
            elif fecha_valor:
                try:
                    fecha = datetime.strptime(str(fecha_valor), '%Y-%m-%d').date()
                except ValueError:
                    continue
            else:
                continue

            precio_unitario = Decimal(str(registro.get('Precio', 0) or 0))
            cantidad        = Decimal(str(registro.get('Cantidad', 0) or 0))
            importe         = Decimal(str(registro.get('Ventas', 0) or 0))
            subtotal        = importe
            impuesto        = Decimal('0')
            total           = importe

            categoria, _ = Categoria.objects.get_or_create(nombre=categoria_nombre)

            sucursal, _ = Sucursal.objects.get_or_create(
                nombre=empresa_nombre,
                defaults={
                    'ciudad': ciudad,
                    'estado': provincia,
                    'pais': 'Ecuador'
                }
            )

            metodo_pago, _ = MetodoPago.objects.get_or_create(nombre=metodo_nombre)
            cliente, _     = Cliente.objects.get_or_create(nombre=cliente_nombre)
            vendedor, _    = Vendedor.objects.get_or_create(
                nombre=vendedor_nombre,
                defaults={'sucursal': sucursal}
            )

            producto, _ = Producto.objects.get_or_create(
                codigo=nombre_producto[:50],
                defaults={
                    'nombre': nombre_producto,
                    'categoria': categoria,
                    'precio': precio_unitario
                }
            )

            folio_str = str(folio)
            if folio_str not in ventas_temporales:
                ventas_temporales[folio_str] = {
                    'fecha': fecha,
                    'cliente': cliente,
                    'sucursal': sucursal,
                    'vendedor': vendedor,
                    'metodo_pago': metodo_pago,
                    'subtotal': subtotal,
                    'impuesto': impuesto,
                    'total': total,
                    'detalles': []
                }
            else:
                ventas_temporales[folio_str]['total']    += importe
                ventas_temporales[folio_str]['subtotal'] += importe

            ventas_temporales[folio_str]['detalles'].append({
                'producto': producto,
                'cantidad': cantidad,
                'precio_unitario': precio_unitario,
                'importe': importe
            })

        for folio, info in ventas_temporales.items():
            venta, creada = Venta.objects.get_or_create(
                folio=folio,
                defaults={
                    'fecha':       info['fecha'],
                    'cliente':     info['cliente'],
                    'sucursal':    info['sucursal'],
                    'vendedor':    info['vendedor'],
                    'metodo_pago': info['metodo_pago'],
                    'subtotal':    info['subtotal'],
                    'impuesto':    info['impuesto'],
                    'total':       info['total'],
                }
            )

            if creada:
                for d in info['detalles']:
                    DetalleVenta.objects.create(
                        venta=venta,
                        producto=d['producto'],
                        cantidad=d['cantidad'],
                        precio_unitario=d['precio_unitario'],
                        importe=d['importe']
                    )

        self.stdout.write(self.style.SUCCESS('Importación completada correctamente'))python manage.py importar_basededatos --archivo "DashBoard2021 (1).xlsx"
